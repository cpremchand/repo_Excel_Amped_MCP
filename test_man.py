"""
Excel Test Case MCP Server (Professional, Procedural, AI/Agent Ready)

- Manages SW Validation Testing sheets with proper test case appending
- Thread-safe operations for multi-user scenarios
- Template-based Excel document creation and manipulation
- Systematic test case management with validation
"""

from mcp.server.fastmcp import FastMCP
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from typing import List, Dict, Any, Optional
import os
import threading
from datetime import datetime
import textwrap
import re
from typing import Optional


mcp = FastMCP("Excel Test Case MCP Server")

'''
# --- Helper Functions ---
def get_next_available_row(ws, sheet_name: str = "SW Validation Testing") -> int:
    ...
    return row
'''


def auto_wrap_text(text: str, max_len: int = 50) -> str:
    """
    Breaks long text into multiple lines at word boundaries
    to help Excel auto-expand row height.
    """
    if not text:
        return ""
    return '\n'.join(textwrap.wrap(text, width=max_len))

def step_wrap(text: str) -> str:
    """
    Adds newlines before each numbered step (1., 2., 3., ...) for clearer formatting.
    """
    # Clean up input whitespace
    text = text.strip()

    # Ensure new line before each number + dot (e.g., 2. or 10.)
    return re.sub(r'\s*(?=\d+\.)', '\n', text)


# --- Thread-safe Workbook Store ---
class WorkbookStore:
    def __init__(self):
        self._workbooks = {}
        self._lock = threading.Lock()
    
    def create_from_template(self, wb_id: str, template_path: str):
        with self._lock:
            if wb_id in self._workbooks:
                raise ValueError(f"Workbook '{wb_id}' already exists.")
            if not os.path.exists(template_path):
                raise ValueError(f"Template '{template_path}' does not exist.")
            self._workbooks[wb_id] = load_workbook(template_path)
    
    def create_blank(self, wb_id: str):
        with self._lock:
            if wb_id in self._workbooks:
                raise ValueError(f"Workbook '{wb_id}' already exist.")
            self._workbooks[wb_id] = Workbook()
    
    def get(self, wb_id: str):
        with self._lock:
            if wb_id not in self._workbooks:
                raise ValueError(f"Workbook '{wb_id}' does not exist.")
            return self._workbooks[wb_id]
    
    def delete(self, wb_id: str):
        with self._lock:
            if wb_id in self._workbooks:
                del self._workbooks[wb_id]
    
    def exists(self, wb_id: str) -> bool:
        with self._lock:
            return wb_id in self._workbooks

WORKBOOKS = WorkbookStore()

# --- Configuration ---
CONFIG = {
    "default_template_path": os.path.join(os.path.dirname(__file__), "sample.xlsx"),
    "test_case_start_row": 13,
    "max_test_cases": 100,  # Rows 13 to 112 (100 test cases)
    "header_row": 12,
}

# --- Styling Constants ---
STYLES = {
    "header_fill": PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"),
    "bold_font": Font(bold=True),
    "title_font": Font(bold=True, size=14),
    "italic_font": Font(italic=True),
    "center_wrap": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "left_wrap": Alignment(wrap_text=True, vertical="center", horizontal="left"),
    "black_border": Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
}

# --- Test Case Schema ---
TEST_CASE_COLUMNS = [
    "Traceability Req-ID",      # B (2)
    "Test Case ID",             # C (3) 
    "Priority",                 # D (4)
    "Test Case Objective",      # E (5)
    "Test Precondition",        # F (6)
    "Test Steps",               # G (7)
    "Test Inputs",              # H (8)
    "Test Case Design Methodology", # I (9)
    "Dependent Test Cases",     # J (10)
    "Expected Outcome",         # K (11)
    "Actual Outcome",           # L (12)
    "Test Result",              # M (13)
    "Remarks",                  # N (14)
    "Track Bug ID"              # O (15)
]

# --- Helper Functions ---
def get_next_available_row(ws, sheet_name: str = "SW Validation Testing") -> int:
    """Find the next available row for adding test cases"""
    if ws.title != sheet_name:
        return CONFIG["test_case_start_row"]
    
    start_row = CONFIG["test_case_start_row"]
    max_row = start_row + CONFIG["max_test_cases"] - 1
    
    # Check from start row to find first empty row
    for row in range(start_row, max_row + 1):
        # Check if Test Case ID column (C) is empty
        if ws.cell(row=row, column=3).value is None or str(ws.cell(row=row, column=3).value).strip() == "":
            return row
    
    # If all rows are filled, return next row (might exceed template)
    return max_row + 1

def setup_sw_validation_sheet(ws):
    """Create the complete SW Validation Testing sheet structure"""
    ws.title = "SW Validation Testing"
    
    # Set column widths
    col_widths = {
        'B': 25, 'C': 15, 'D': 10, 'E': 25, 'F': 25, 'G': 25, 'H': 25,
        'I': 25, 'J': 25, 'K': 25, 'L': 25, 'M': 15, 'N': 30, 'O': 15, 'P': 15
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # Row 2: Main title
    ws.merge_cells('B2:P2')
    ws['B2'] = "3.0 Software Validation Testing"
    ws['B2'].font = STYLES["title_font"]
    ws['B2'].alignment = STYLES["center_wrap"]
    
    # Row 4: Section title
    ws['B4'] = "3.1 Testing Details"
    ws['B4'].font = STYLES["bold_font"]
    
    # Testing details section
    testing_details = {
        """Generate the testing details section with appropriate labels and styles"""
        'B5': "Project Name and ID",      'I5': "Test Environment",
        'B6': "Features to be Tested",    'I6': "Test Case Designer",
        'B7': "References/Input Documents with Version", 'I7': "Test Case Reviewer",
        'B8': "Common Attributes",        'I8': "Tester",
        'B9': "Notation for description", 'I9': "Test Start Date", 'J9': "dd-Mmm-yyyy",
        'B10': "Version of Item under test", 'I10': "Test End Date", 'J10': "dd-Mmm-yyyy"
    }
    for cell, value in testing_details.items():
        ws[cell] = value
        if cell[0] in ['B', 'I']:
            ws[cell].font = STYLES["bold_font"]
    
    # Dynamic row range
    start_row = CONFIG["test_case_start_row"]
    end_row = start_row + CONFIG["max_test_cases"] - 1

    # Total Test Cases count (based on non-empty Test Case ID column)
    ws['N9'] = "Total Test Cases"
    ws['N9'].font = STYLES["bold_font"]
    ws['N9'].alignment = STYLES["center_wrap"]
    ws['N9'].border = STYLES["black_border"]
    ws.merge_cells('N9:N10')

    ws['O9'] = f'=COUNTA(C{start_row}:C{end_row})'
    ws['O9'].alignment = STYLES["center_wrap"]
    ws['O9'].border = STYLES["black_border"]
    ws.merge_cells('O9:O10')

    summary_data = [
        ("Total No of Bugs Identified", f'=COUNTA(O{start_row}:O{end_row})'),
        ("Passed Test Cases", f'=COUNTIF(M{start_row}:M{end_row},"Passed")'),
        ("Failed Test Cases", f'=COUNTIF(M{start_row}:M{end_row},"Failed")'),
        ("Test Cases Not Tested", f'=COUNTIF(M{start_row}:M{end_row},"Not Tested")')
    ]

    for i, (label, formula) in enumerate(summary_data):
        row = 5 + i
        ws[f'N{row}'] = label
        ws[f'N{row}'].font = STYLES["bold_font"]
        ws[f'N{row}'].alignment = STYLES["center_wrap"]
        ws[f'N{row}'].border = STYLES["black_border"]
        
        ws[f'O{row}'] = formula
        ws[f'O{row}'].alignment = STYLES["center_wrap"]
        ws[f'O{row}'].border = STYLES["black_border"]
    
    # Section title for results
    ws['B11'] = "3.2 Validation Testing Results"
    ws['B11'].font = STYLES["bold_font"]
    
    # Table headers
    headers = [
        "Traceability\nReq-ID", "Test Case\nID", "Priority", "Test Case Objective",
        "Test Precondition", "Test Steps", "Test Inputs\n(Conditions / Values)",
        "Test Case Design\nMethodology", "Dependent\nTest Cases",
        "Expected\nOutcome", "Actual\nOutcome", "Test Result", "Remarks", "Track Bug ID\n(If Applicable)"
    ]
    
    for col_index, header in enumerate(headers, start=2):
        col_letter = get_column_letter(col_index)
        cell = ws[f'{col_letter}12']
        cell.value = header
        cell.font = STYLES["bold_font"]
        cell.alignment = STYLES["center_wrap"]
        cell.fill = STYLES["header_fill"]
        cell.border = STYLES["black_border"]
    

    for row in range(start_row, end_row + 1):
        for col in range(2, 16):  # Columns B to O
            ws.cell(row=row, column=col).border = STYLES["black_border"]
            
    for row in range(start_row, end_row + 1):
        ws[f"M{row}"].value = ""  # Force-create the cell

    
    # Add dropdown validation for Test Result column
    dv = DataValidation(type="list", formula1='"Not Tested,Passed,Failed"', allow_blank=True)
    dv.error = 'Select a value from the list'
    dv.errorTitle = 'Invalid Input'
    dv.prompt = 'Please select a test result'
    dv.promptTitle = 'Test Result Selection'
    ws.add_data_validation(dv)
    dv.add(f"M{start_row}:M{end_row}")  #  Dynamic range

# --- MCP Tools ---

@mcp.tool(description="Create a new Excel workbook from template. If template_path not provided, creates SW Validation Testing sheet from scratch.")
def create_workbook(wb_id: str, template_path: Optional[str] = None) -> str:
    try:
        if template_path is None:
            template_path = CONFIG.get("default_template_path")
        
        if template_path and os.path.exists(template_path):
            WORKBOOKS.create_from_template(wb_id, template_path)
            return f"Workbook '{wb_id}' created from template '{template_path}'."
        else:
            WORKBOOKS.create_blank(wb_id)
            wb = WORKBOOKS.get(wb_id)
            setup_sw_validation_sheet(wb.active)
            if template_path:
                return f"Workbook '{wb_id}' created with SW Validation sheet (template '{template_path}' not found)."
            else:
                return f"Workbook '{wb_id}' created with SW Validation Testing sheet."
    except Exception as e:
        return f"Error: {str(e)}"

@mcp.tool(description="Open an existing Excel workbook from file path.")
def open_workbook(wb_id: str, filepath: str) -> str:
    try:
        if WORKBOOKS.exists(wb_id):
            return f"Error: Workbook '{wb_id}' already exists in memory."
        
        if not os.path.exists(filepath):
            return f"Error: File '{filepath}' does not exist."
        
        WORKBOOKS.create_from_template(wb_id, filepath)
        return f"Workbook '{wb_id}' opened from '{filepath}'."
    except Exception as e:
        return f"Error: {str(e)}"

@mcp.tool(description="Add a single test case to the SW Validation Testing sheet, SW Intergration Testing Sheet, SW Unit Testing Sheet.")
def add_test_case(
    wb_id: str,
    traceability_req_id: str,
    test_case_id: str,
    priority: str,
    test_case_objective: str,
    test_precondition: str,
    test_steps: str,
    test_inputs: str,
    test_case_design_methodology: str,
    dependent_test_cases: str = "None",
    expected_outcome: str = "",
    actual_outcome: str = "",
    test_result: str = "Not Tested",
    remarks: str = "",
    track_bug_id: str = "",
    sheet_name: str = "SW Validation Testing"
) -> str:
    try:
        wb = WORKBOOKS.get(wb_id)

        # Get the target worksheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            return f"Error: Sheet '{sheet_name}' not found in workbook '{wb_id}'."

        # Find next available row
        next_row = get_next_available_row(ws, sheet_name)

        # Validate test result
        valid_results = ["Not Tested", "Passed", "Failed"]
        if test_result not in valid_results:
            test_result = "Not Tested"

        # Prepare test case data
        test_case_data = [
            auto_wrap_text(traceability_req_id),
            auto_wrap_text(test_case_id),
            auto_wrap_text(priority),
            auto_wrap_text(test_case_objective),
            auto_wrap_text(test_precondition),
            step_wrap(test_steps),
            auto_wrap_text(test_inputs),
            auto_wrap_text(test_case_design_methodology),
            auto_wrap_text(dependent_test_cases),
            auto_wrap_text(expected_outcome),
            auto_wrap_text(actual_outcome),
            auto_wrap_text(test_result),
            auto_wrap_text(remarks),
            auto_wrap_text(track_bug_id)
        ]

        # Write test case data to worksheet
        max_lines = 1
        for col_index, value in enumerate(test_case_data, start=2):
            cell = ws.cell(row=next_row, column=col_index)
            cell.value = value
            # Apply alignment with proper wrapping
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            cell.border = STYLES["black_border"]
            cell.font = Font(color="000000", size=9)
            if isinstance(value, str):
                lines = value.count('\n') + 1
                max_lines = max(max_lines, lines)

        # Set row height based on most wrapped lines
        ws.row_dimensions[next_row].height = max(15, max_lines * 15)

        return f"Test case '{test_case_id}' added to row {next_row} in sheet '{sheet_name}'."
    except Exception as e:
        return f"Error: {str(e)}"


@mcp.tool(description="Add multiple test cases at once. Each test case should be a dictionary with required fields.")
def add_multiple_test_cases(wb_id: str, test_cases: List[Dict[str, str]], sheet_name: str = "SW Validation Testing") -> str:
    try:
        wb = WORKBOOKS.get(wb_id)

        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found in workbook '{wb_id}'."

        ws = wb[sheet_name]
        added_count = 0

        for test_case in test_cases:
            try:
                # Extract required fields with defaults
                traceability_req_id = auto_wrap_text(test_case.get("traceability_req_id", ""))
                test_case_id = auto_wrap_text(test_case.get("test_case_id", ""))
                priority = auto_wrap_text(test_case.get("priority", "Medium"))
                test_case_objective = auto_wrap_text(test_case.get("test_case_objective", ""))
                test_precondition = auto_wrap_text(test_case.get("test_precondition", ""))
                test_steps = step_wrap(test_case.get("test_steps", ""))
                test_inputs = auto_wrap_text(test_case.get("test_inputs", ""))
                test_case_design_methodology = auto_wrap_text(test_case.get("test_case_design_methodology", ""))
                dependent_test_cases = auto_wrap_text(test_case.get("dependent_test_cases", "None"))
                expected_outcome = auto_wrap_text(test_case.get("expected_outcome", ""))
                actual_outcome = auto_wrap_text(test_case.get("actual_outcome", ""))
                test_result = auto_wrap_text(test_case.get("test_result", "Not Tested"))
                remarks = auto_wrap_text(test_case.get("remarks", ""))
                track_bug_id = auto_wrap_text(test_case.get("track_bug_id", ""))

                # Find next available row
                next_row = get_next_available_row(ws, sheet_name)

                # Validate test result
                valid_results = ["Not Tested", "Passed", "Failed"]
                if test_result not in valid_results:
                    test_result = "Not Tested"

                # Prepare and write test case data
                test_case_data = [
                    traceability_req_id, test_case_id, priority, test_case_objective,
                    test_precondition, test_steps, test_inputs, test_case_design_methodology,
                    dependent_test_cases, expected_outcome, actual_outcome, test_result,
                    remarks, track_bug_id
                ]

                max_lines = 1
                for col_index, value in enumerate(test_case_data, start=2):
                    cell = ws.cell(row=next_row, column=col_index)
                    cell.value = value
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                    cell.border = STYLES["black_border"]
                    cell.font = Font(color="000000", size=9)
                    if isinstance(value, str):
                        lines = value.count('\n') + 1
                        max_lines = max(max_lines, lines)

                ws.row_dimensions[next_row].height = max(15, max_lines * 15)
                added_count += 1

            except Exception as tc_error:
                print(f"Warning: Failed to add test case {test_case.get('test_case_id', 'Unknown')}: {str(tc_error)}")

        return f"Successfully added {added_count} test cases to sheet '{sheet_name}'."
    except Exception as e:
        return f"Error: {str(e)}"


@mcp.tool(description="Update a specific test case by Test Case ID.")
def update_test_case(
    wb_id: str,
    test_case_id: str,
    field_updates: Dict[str, str],
    sheet_name: str = "SW Validation Testing"
) -> str:
    try:
        wb = WORKBOOKS.get(wb_id)
        
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found."
        
        ws = wb[sheet_name]
        
        # Find the test case by ID (column C)
        test_case_row = None
        start_row = CONFIG["test_case_start_row"]
        end_row = start_row + CONFIG["max_test_cases"]
        
        for row in range(start_row, end_row):
            if ws.cell(row=row, column=3).value == test_case_id:
                test_case_row = row
                break
        
        if test_case_row is None:
            return f"Error: Test case '{test_case_id}' not found."
        
        # Map field names to column indices
        field_to_col = {
            "traceability_req_id": 2, "test_case_id": 3, "priority": 4,
            "test_case_objective": 5, "test_precondition": 6, "test_steps": 7,
            "test_inputs": 8, "test_case_design_methodology": 9,
            "dependent_test_cases": 10, "expected_outcome": 11,
            "actual_outcome": 12, "test_result": 13, "remarks": 14,
            "track_bug_id": 15
        }
        
        updated_fields = []
        for field_name, new_value in field_updates.items():
            if field_name in field_to_col:
                col_index = field_to_col[field_name]
                
                # Validate test result
                if field_name == "test_result" and new_value not in ["Not Tested", "Passed", "Failed"]:
                    new_value = "Not Tested"
                
                ws.cell(row=test_case_row, column=col_index).value = new_value
                updated_fields.append(field_name)
        
        return f"Updated test case '{test_case_id}' fields: {', '.join(updated_fields)}."
    except Exception as e:
        return f"Error: {str(e)}"

@mcp.tool(description="Get all test cases from the SW Validation Testing sheet.")
def get_all_test_cases(wb_id: str, sheet_name: str = "SW Validation Testing") -> str:
    try:
        wb = WORKBOOKS.get(wb_id)
        
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found."
        
        ws = wb[sheet_name]
        test_cases = []
        
        start_row = CONFIG["test_case_start_row"]
        end_row = start_row + CONFIG["max_test_cases"]
        
        for row in range(start_row, end_row):
            # Check if Test Case ID is not empty
            test_case_id = ws.cell(row=row, column=3).value
            if test_case_id and str(test_case_id).strip():
                test_case = {}
                for col_index, field_name in enumerate(TEST_CASE_COLUMNS, start=2):
                    cell_value = ws.cell(row=row, column=col_index).value
                    test_case[field_name] = cell_value if cell_value is not None else ""
                test_cases.append(f"Row {row}: {test_case}")
        
        if not test_cases:
            return f"No test cases found in sheet '{sheet_name}'."
        
        return f"Found {len(test_cases)} test cases:\n" + "\n".join(test_cases)
    except Exception as e:
        return f"Error: {str(e)}"

@mcp.tool(description="Get test case statistics and summary.")
def get_test_case_summary(wb_id: str, sheet_name: str = "SW Validation Testing") -> str:
    try:
        wb = WORKBOOKS.get(wb_id)
        
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found."
        
        ws = wb[sheet_name]
        
        total_count = 0
        passed_count = 0
        failed_count = 0
        not_tested_count = 0
        bug_count = 0
        
        start_row = CONFIG["test_case_start_row"]
        end_row = start_row + CONFIG["max_test_cases"]
        
        for row in range(start_row, end_row):
            # Check if Test Case ID is not empty
            test_case_id = ws.cell(row=row, column=3).value
            if test_case_id and str(test_case_id).strip():
                total_count += 1
                
                # Check test result (column M = 13)
                test_result = ws.cell(row=row, column=13).value
                if test_result == "Passed":
                    passed_count += 1
                elif test_result == "Failed":
                    failed_count += 1
                else:
                    not_tested_count += 1
                
                # Check for bug ID (column O = 15)
                bug_id = ws.cell(row=row, column=15).value
                if bug_id and str(bug_id).strip():
                    bug_count += 1
        
        summary = f"""Test Case Summary for '{sheet_name}':
Total Test Cases: {total_count}
Passed: {passed_count}
Failed: {failed_count}
Not Tested: {not_tested_count}
Test Cases with Bugs: {bug_count}
Next Available Row: {get_next_available_row(ws, sheet_name)}"""
        
        return summary
    except Exception as e:
        return f"Error: {str(e)}"

@mcp.tool(description="Update testing details section (project info, dates, etc.) also for SW Intergration Testing Sheet and SW Unit Testing Sheet.")
def update_testing_details(
    wb_id: str,
    project_name: str = "",
    features_to_test: str = "",
    references: str = "",
    common_attributes: str = "",
    notation: str = "",
    version_under_test: str = "",
    test_environment: str = "",
    test_case_designer: str = "",
    test_case_reviewer: str = "",
    tester: str = "",
    test_start_date: str = "",
    test_end_date: str = "",
    sheet_name: str = "SW Validation Testing"
) -> str:
    try:
        wb = WORKBOOKS.get(wb_id)
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found."
 
        ws = wb[sheet_name]
 
        # Mapping for left block (Project details: B5:D10 label, E5:H10 value)
        left_labels = [
            "Project Name and ID",
            "Features to be Tested",
            "References/Input Documents with Version",
            "Common Attributes",
            "Notation for description",
            "Version of Item under test"
        ]
        left_values = [
            project_name,
            features_to_test,
            references,
            common_attributes,
            notation,
            version_under_test
        ]
 
        # Mapping for right block (Test details: I5:I10 label, J5:M10 value)
        right_labels = [
            "Test Environment",
            "Test Case Designer",
            "Test Case Reviewer",
            "Tester",
            "Test Start Date",
            "Test End Date"
        ]
        right_values = [
            test_environment,
            test_case_designer,
            test_case_reviewer,
            tester,
            test_start_date,
            test_end_date
        ]
 
        # Fill left block (B5:D10 label, E5:H10 value)
        for i, (label, value) in enumerate(zip(left_labels, left_values)):
            row = 5 + i
            cell = ws.cell(row=row, column=5, value=value)
            cell.alignment = STYLES["left_wrap"]
            cell.border = STYLES["black_border"]
            cell.font = Font(color="000000", size=9)
 
        # Fill right block (I5:I10 label, J5:M10 value)
        for i, (label, value) in enumerate(zip(right_labels, right_values)):
            row = 5 + i
            cell = ws.cell(row=row, column=10, value=value)
            cell.alignment = STYLES["left_wrap"]
            cell.border = STYLES["black_border"]
            cell.font = Font(color="000000", size=9)
 
        return f"Testing details updated on sheet '{sheet_name}'."
    except Exception as e:
        return f"Error: {str(e)}"


@mcp.tool(description="Extract testing details from the given sheet (SW Validation Testing, SW Intergration Testing, SW Unit Testing).")
def get_testing_details(wb_id: str, sheet_name: str = "SW Validation Testing") -> dict:
    try:
        wb = WORKBOOKS.get(wb_id)
        if sheet_name not in wb.sheetnames:
            return {"error": f"Sheet '{sheet_name}' not found in workbook."}
        
        ws = wb[sheet_name]

        details = {
            "Project Name and ID": ws["E5"].value,
            "Features to be Tested": ws["E6"].value,
            "References": ws["E7"].value,
            "Common Attributes": ws["E8"].value,
            "Notation": ws["E9"].value,
            "Version under Test": ws["E10"].value,
            "Test Environment": ws["J5"].value,
            "Test Case Designer": ws["J6"].value,
            "Test Case Reviewer": ws["J7"].value,
            "Tester": ws["J8"].value,
            "Test Start Date": ws["J9"].value,
            "Test End Date": ws["J10"].value,
        }

        return details

    except Exception as e:
        return {"error": str(e)}


@mcp.tool(description="Save the workbook to a .xlsx file.")
def save_workbook(wb_id: str, filepath: str) -> str:
    try:
        wb = WORKBOOKS.get(wb_id)
        wb.save(filepath)
        return f"Workbook '{wb_id}' saved to {filepath}."
    except Exception as e:
        return f"Error: {str(e)}"

@mcp.tool(description="Get workbook basic information.")
def get_workbook_info(wb_id: str) -> str:
    try:
        wb = WORKBOOKS.get(wb_id)
        
        info = f"Workbook '{wb_id}' information:\n"
        info += f"  Sheets: {', '.join(wb.sheetnames)}\n"
        
        # If SW Validation Testing sheet exists, provide test case count
        if "SW Validation Testing" in wb.sheetnames:
            ws = wb["SW Validation Testing"]
            test_count = 0
            start_row = CONFIG["test_case_start_row"]
            end_row = start_row + CONFIG["max_test_cases"]
            
            for row in range(start_row, end_row):
                test_case_id = ws.cell(row=row, column=3).value
                if test_case_id and str(test_case_id).strip():
                    test_count += 1
            
            info += f"  Test Cases in SW Validation Testing: {test_count}\n"
            info += f"  Next Available Row: {get_next_available_row(ws)}"
        
        return info
    except Exception as e:
        return f"Error: {str(e)}"

@mcp.tool(description="Delete a workbook from memory.")
def delete_workbook(wb_id: str) -> str:
    try:
        if not WORKBOOKS.exists(wb_id):
            return f"Error: Workbook '{wb_id}' does not exist."
        
        WORKBOOKS.delete(wb_id)
        return f"Workbook '{wb_id}' deleted from memory."
    except Exception as e:
        return f"Error: {str(e)}"

"""
@mcp.tool(description="Get all requirements from the Functional Requirements sheet and Non-Functional Requirements sheet.")
def get_requirments_from_srs_xlsx_tool(wb_id: str, sheet_name: str = "Functional Requirements and Non-Functional Requirements") -> str:
    try:
        wb = WORKBOOKS.get(wb_id)
        
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found."
        
        ws = wb[sheet_name]
        test_cases = []
        
        start_row = CONFIG["test_case_start_row"]
        end_row = start_row + CONFIG["max_test_cases"]
        
        for row in range(start_row, end_row):
            # Check if Test Case ID is not empty
            test_case_id = ws.cell(row=row, column=3).value
            if test_case_id and str(test_case_id).strip():
                test_case = {}
                for col_index, field_name in enumerate(TEST_CASE_COLUMNS, start=2):
                    cell_value = ws.cell(row=row, column=col_index).value
                    test_case[field_name] = cell_value if cell_value is not None else ""
                test_cases.append(f"Row {row}: {test_case}")
        
        if not test_cases:
            return f"No test cases found in sheet '{sheet_name}'."
        
        return f"Found {len(test_cases)} test cases:\n" + "\n".join(test_cases)
    except Exception as e:
        return f"Error: {str(e)}"

"""
if __name__ == "__main__":
    mcp.run()
